"""
FRIDAY-OUTREACH
Full sales pipeline: find leads → enrich with real site/email data → export spreadsheet → email.

Pipeline:
  Google Places API → seen.py dedup → sitecheck.py → scrape.py → openpyxl Excel → Brevo email
"""

import os
import re
import sys
import json
import asyncio
import requests
from datetime import datetime
from .base import BaseAgent

# ── Central config (cloud-ready paths) ───────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from config import LEADS_DIR, COMPANY_NAME, COMPANY_EMAIL, REPLY_TO
import state as _state

# ── LeadForge modules ─────────────────────────────────────────────────────────
_LF = os.path.expanduser("~/Documents/leadforge/app")
if _LF not in sys.path:
    sys.path.insert(0, _LF)

try:
    from leadforge.core import sitecheck as _sitecheck
    from leadforge.core import scrape    as _scrape
    from leadforge.core import seen      as _seen
    _LF_AVAILABLE = True
except ImportError:
    _LF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# ── Config ────────────────────────────────────────────────────────────────────
BREVO_API_KEY = os.getenv("BREVO_API_KEY", "")

SIGNAL_SENTENCES = {
    "no website":            "No website at all — easiest pitch, huge opportunity",
    "website unreachable":   "Website is down or broken — perfect opening",
    "basic / thin site":     "Thin, bare-bones site — big room for improvement",
    "not mobile optimized":  "Site not mobile-friendly — Google penalises this",
    "template-based website":"Cookie-cutter template site (Wix/WP) — easy upsell to custom build",
}


class FridayOutreach(BaseAgent):
    provider = "anthropic"
    model    = "claude-sonnet-4-6"

    # ── In-process cache (persistent store is state.py) ──────────────────────
    _user_id: str = "web"   # overridden per Telegram user in handle()

    system_prompt = """You are FRIDAY, the outreach agent for FileFlow Inc. You handle cold emails only.

CRITICAL RULE — READ THIS FIRST:
You do NOT find leads. You do NOT know any real businesses. You do NOT have internet access.
NEVER invent, list, suggest, or make up any business names, websites, phone numbers, or contacts.
If asked to find leads or businesses, say ONLY: "Use the Outreach module and say: find [niche] in [city]"

YOUR ONLY JOB: Write and send cold emails when given specific context.

Cold email rules:
- Subject: under 8 words, no spam triggers
- First line: specific to their business type
- Body: 4-5 sentences max
- CTA: "worth a quick call this week?"
- Human tone, never corporate

Output format when writing emails:
SUBJECT: [subject]
---
[email body only — nothing else]
---"""

    # ── Main handler ──────────────────────────────────────────────────────────

    async def handle(self, message: str, update=None, image_url: str = None, history=None) -> str:
        # ── Resolve user ID for persistent state ─────────────────────────────
        uid = str(update.effective_user.id) if update and hasattr(update, "effective_user") and update.effective_user else "web"
        ctx = _state.get_context(uid)

        last_niche    = ctx.get("last_niche",    "")
        last_location = ctx.get("last_location", "")

        msg_lower = message.lower().strip()

        # ── CRM COMMANDS ──────────────────────────────────────────────────────
        crm_reply = self._handle_crm(msg_lower, message)
        if crm_reply is not None:
            return crm_reply

        # ── PIPELINE VIEW ─────────────────────────────────────────────────────
        if any(x in msg_lower for x in ["show pipeline", "my pipeline", "pipeline status",
                                         "who to follow", "follow up", "followup", "follow-up"]):
            return self._pipeline_report()

        # ── SEND MODE ─────────────────────────────────────────────────────────
        has_email = "@" in message
        is_send = (
            "send to" in msg_lower or "send email to" in msg_lower or
            "send it to" in msg_lower or "email to" in msg_lower or
            (("send" in msg_lower or "email" in msg_lower) and has_email)
        )

        if is_send and has_email:
            emails = re.findall(r'[\w.\-]+@[\w.\-]+\.\w+', message)
            if emails:
                to_email = emails[0]
                ctx_str  = re.sub(
                    r'(send(\s+email)?\s+(it\s+)?to|email(\s+to)?)\s+[\w.\-]+@[\w.\-]+\.\w+',
                    '', message, flags=re.IGNORECASE
                ).strip() or "a local business that needs a professional website"

                draft         = await self._generate_clean_email(ctx_str, to_email)
                subject, body = self._parse_email(draft)
                result        = self._send_via_brevo(to_email, subject, body)

                return (
                    f"Sent to {to_email}\n\n"
                    f"Subject: {subject}\n"
                    f"{'─' * 40}\n"
                    f"{body}\n"
                    f"{'─' * 40}\n\n"
                    f"{result}"
                )

        # ── LEAD FIND MODE ────────────────────────────────────────────────────
        niche, location, count = self._parse_lead_request(message, last_niche, last_location)
        if niche and location:
            _state.set_context(uid, last_niche=niche, last_location=location)
            return await self._find_leads(niche, location, count)

        # ── GUARD: sounds like a lead request but parsing failed ──────────────
        count_only    = re.match(r'^\s*(\d+)\s*$', msg_lower.strip())
        sounds_like   = (
            any(w in msg_lower for w in self.BUSINESS_WORDS) and
            ("in " in msg_lower or "on " in msg_lower or
             "find" in msg_lower or "lead" in msg_lower)
        )

        if count_only or re.match(r'^\s*\d+\s+(more|again)?\s*$', msg_lower):
            if last_niche and last_location:
                n = int(re.search(r'\d+', message).group())
                n = min(max(n, 5), 60)
                return await self._find_leads(last_niche, last_location, n)

        if sounds_like:
            if last_location and not re.search(r'\b(?:in|on)\b', msg_lower):
                for bword in self.BUSINESS_WORDS:
                    if bword in msg_lower:
                        count_m = re.search(r'\b(\d+)\b', msg_lower)
                        count   = int(count_m.group(1)) if count_m else 20
                        count   = min(max(count, 5), 60)
                        _state.set_context(uid, last_niche=bword)
                        return await self._find_leads(bword, last_location, count)

            return (
                "Couldn't parse the niche and city from that, Boss.\n\n"
                "Try something like:\n"
                "  • Find 15 dentists in Kansas City\n"
                "  • 20 restaurants in Miami\n"
                "  • Plumbers on Long Island\n\n"
                "I'll pull real Google Places data — no guessing."
            )

        return await super().handle(message, update, image_url, history=history)

    # ── Lead request parser ───────────────────────────────────────────────────

    BUSINESS_WORDS = [
        "restaurant","restaurants","barber","barbers","salon","salons","dentist","dentists",
        "dental","plumber","plumbers","gym","gyms","lawyer","lawyers","cafe","cafes",
        "hotel","hotels","shop","shops","store","stores","clinic","clinics","agency",
        "agencies","contractor","contractors","spa","spas","photographer","photographers",
        "realtor","realtors","mechanic","mechanics","florist","florists","bakery",
        "bakeries","school","schools","accountant","accountants","therapist","therapists",
        "chiropractor","chiropractors","vet","vets","optician","opticians","doctor",
        "doctors","pharmacy","pharmacies","pizza","burger","sushi","nail","nails",
        "tattoo","tattoos","yoga","studio","studios","law firm","law firms",
        "roofing","roofer","roofers","hvac","electrician","electricians","cleaner",
        "cleaners","cleaning","landscaper","landscapers","landscaping","pest control",
        "catering","caterer","caterers","daycare","daycare","vet clinic","vet clinics",
    ]

    def _parse_lead_request(self, message: str, last_niche: str = "", last_location: str = ""):
        msg = message.lower()

        lead_triggers = ["find","lead","prospect","business","client","get","show",
                         "give","search","look for","need","want","looking for"]
        has_trigger   = any(w in msg for w in lead_triggers + self.BUSINESS_WORDS)
        if not has_trigger:
            return None, None, 20

        # Extract count from anywhere in the message
        count_match = re.search(r'\b(\d+)\b', msg)
        count = int(count_match.group(1)) if count_match else 20
        count = min(max(count, 5), 60)

        # Primary pattern: "X in|on Y" — stops at natural boundaries
        boundary = (
            r'(?=\s*[,;.]|\s+so\b|\s+that\b|\s+to\b|\s+for\b|\s+because\b'
            r'|\s+and\b|\s+i\b|\s+they\b|\s+who\b|\s+if\b|$)'
        )
        pattern = re.search(
            r'(\d+\s+)?([a-z][a-z\s\-]{1,35}?)\s+(?:in|on)\s+([a-z][a-z\s,]{1,40}?)' + boundary,
            msg
        )

        if pattern:
            niche    = pattern.group(2).strip()
            location = pattern.group(3).strip()

            # Clean up niche — remove filler words
            niche = re.sub(r'^(find|me|some|us|the|all|any|a|an)\s+', '', niche).strip()
            niche = re.sub(r'\s+(clinic|business|businesses|place|places|near|around|area)$', '', niche).strip()

            # Clean up location — remove trailing filler
            location = re.sub(r'\s+(area|city|state|county|region)$', '', location).strip()
            location = location.rstrip(',. ')

            if len(niche) >= 2 and len(location) >= 2:
                return niche, location, count

        # Fallback: business word + location after "in" or "on"
        for bword in self.BUSINESS_WORDS:
            if bword in msg:
                loc_match = re.search(
                    r'\b(?:in|on)\s+([A-Za-z][A-Za-z\s,]{2,40}?)' + boundary, msg
                )
                if loc_match:
                    location = loc_match.group(1).strip().rstrip(',. ')
                    if len(location) >= 2:
                        return bword, location, count

        # Fallback: use persistent memory if message just has a niche
        if last_location:
            for bword in self.BUSINESS_WORDS:
                if bword in msg:
                    return bword, last_location, count

        return None, None, 20

    # ── Core lead finder ──────────────────────────────────────────────────────

    async def _find_leads(self, niche: str, location: str, count: int) -> str:
        api_key = os.getenv("GOOGLE_PLACES_API_KEY", "")
        if not api_key:
            return "No Google Places API key found. Add GOOGLE_PLACES_API_KEY to config/.env"

        try:
            raw_places = await self._fetch_places(niche, location, count * 2, api_key)
            if not raw_places:
                return f"No results for '{niche} in {location}'. Try a different city or niche."

            # ── Deduplication via seen.py ─────────────────────────────────────
            seen_ids: set = set()
            fresh_places  = raw_places

            if _LF_AVAILABLE:
                try:
                    fresh_places, seen_ids = await asyncio.to_thread(
                        _seen.filter_new, raw_places
                    )
                    if not fresh_places:
                        # All seen — reset and use full set (user wants new run)
                        fresh_places = raw_places
                except Exception:
                    fresh_places = raw_places

            # ── Concurrent enrichment ─────────────────────────────────────────
            sem   = asyncio.Semaphore(5)
            tasks = [self._enrich_place(p, api_key, sem) for p in fresh_places]
            enriched = await asyncio.gather(*tasks)
            leads = [l for l in enriched if l]

            # Sort: hot first (no website), warm next (bad site), cold last (good site)
            leads.sort(key=lambda x: x["sort_key"])
            leads = leads[:count]

            # ── Mark seen ─────────────────────────────────────────────────────
            if _LF_AVAILABLE and leads:
                try:
                    seen_ids.update(l["place_id"] for l in leads if l.get("place_id"))
                    await asyncio.to_thread(_seen.save_seen, seen_ids)
                except Exception:
                    pass

            # ── Save files ────────────────────────────────────────────────────
            os.makedirs(LEADS_DIR, exist_ok=True)
            ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
            slug      = re.sub(r'[^a-z0-9]', '_', niche.lower())[:20]
            city_slug = re.sub(r'[^a-z0-9]', '_', location.lower())[:20]
            base_name = f"leads_{slug}_{city_slug}_{ts}"

            xlsx_path = None
            if XLSX_AVAILABLE:
                xlsx_path = os.path.join(LEADS_DIR, f"{base_name}.xlsx")
                self._make_spreadsheet(leads, niche, location, xlsx_path)

            # JSON hidden backup
            json_path = os.path.join(LEADS_DIR, f".{base_name}.json")
            with open(json_path, 'w') as f:
                json.dump(leads, f, indent=2, ensure_ascii=False)

            # ── Save to persistent pipeline / CRM ─────────────────────────────
            _state.add_leads_to_pipeline(leads, niche, location)

            # ── Build text response ───────────────────────────────────────────
            hot  = sum(1 for l in leads if l["sort_key"] == 0)
            warm = sum(1 for l in leads if l["sort_key"] == 1)
            cold = sum(1 for l in leads if l["sort_key"] == 2)

            lines = [
                f"Found {len(leads)} {niche} leads in {location.title()}",
                f"{hot} hot (no website)  ·  {warm} warm (weak site)  ·  {cold} cold (decent site)\n"
            ]

            for i, lead in enumerate(leads, 1):
                tag = {"Hot": "HOT", "Warm": "WARM", "Cold": "COLD"}.get(lead["priority"], lead["priority"])
                lines.append(f"{i}. [{tag}]  {lead['name']}")
                if lead["phone"]:
                    lines.append(f"   Phone:    {lead['phone']}")
                if lead["email"]:
                    lines.append(f"   Email:    {lead['email']}")
                lines.append(f"   Website:  {lead['website'] or 'None — easiest pitch'}")
                ps = lead.get("perf_score")
                if ps is not None:
                    label = "FAST" if ps >= 80 else ("SLOW" if ps < 50 else "OK")
                    lines.append(f"   Speed:    {ps}/100 ({label})")
                if lead["address"]:
                    lines.append(f"   Address:  {lead['address']}")
                lines.append(f"   Rating:   {lead['rating']}★  ({lead['reviews']} reviews)")
                lines.append(f"   Why:      {lead['why']}")
                lines.append("")

            lines.append("Tell me which ones to email and I'll draft and send it, Boss.")
            lines.append("To update a lead: 'mark lead 3 as contacted' or 'mark Miami Dental as replied'")
            lines.append("To see your pipeline: 'show pipeline'")

            response = "\n".join(lines)

            if xlsx_path and os.path.exists(xlsx_path):
                fname     = os.path.basename(xlsx_path)
                response += f"\n\n[SPREADSHEET:{xlsx_path}|{fname}]"

            return response

        except Exception as e:
            import traceback
            return f"Lead search failed: {str(e)}\n{traceback.format_exc()[-400:]}"

    # ── CRM commands ──────────────────────────────────────────────────────────

    def _handle_crm(self, msg_lower: str, original: str) -> str | None:
        """Handle 'mark lead X as Y' commands. Returns reply or None."""
        # Pattern: "mark lead 3 as contacted" / "mark miami dental as replied"
        m = re.search(
            r'\bmark\b.{0,30}?\b(lead\s+)?(.+?)\s+as\s+(contacted|replied|interested|closed|not.interested)',
            msg_lower
        )
        if not m:
            return None

        identifier = m.group(2).strip()
        status     = m.group(3).replace(" ", "_").replace("-", "_")

        lead = _state.update_lead_status(identifier, status)
        if not lead:
            return f"Couldn't find a lead matching '{identifier}' in your pipeline, Boss."

        name = lead.get("name", identifier)
        msgs = {
            "contacted":     f"Got it — {name} marked as contacted. I'll remind you to follow up in 3 days.",
            "replied":       f"Nice — {name} marked as replied. Follow-up reminder cleared.",
            "interested":    f"Let's go — {name} marked as interested. Time to close.",
            "closed":        f"Closed! {name} is a win. Follow-up cleared.",
            "not_interested": f"Noted — {name} marked as not interested. Moving on.",
        }
        return msgs.get(status, f"{name} updated to {status}.")

    def _pipeline_report(self) -> str:
        from datetime import datetime
        pipeline = _state.get_pipeline()
        summary  = _state.get_pipeline_summary()
        due      = _state.get_followups_due()

        if not pipeline:
            return (
                "No leads in your pipeline yet, Boss.\n\n"
                "Find some first: 'Find 20 dentists in Miami'"
            )

        lines = [
            f"PIPELINE OVERVIEW",
            f"{'─' * 36}",
            f"Total leads:     {summary['total']}",
            f"Not contacted:   {summary['not_contacted']}",
            f"Contacted:       {summary['contacted']}",
            f"Replied:         {summary['replied']}",
            f"Interested:      {summary['interested']}",
            f"Closed (won):    {summary['closed']}",
            f"Not interested:  {summary['not_interested']}",
            "",
        ]

        if due:
            lines.append(f"FOLLOW-UPS DUE ({len(due)})")
            lines.append("─" * 36)
            for lead in due[:10]:
                fu = lead.get("followup_at", "")
                try:
                    days_ago = (datetime.now() - datetime.fromisoformat(fu)).days
                    when = f"{days_ago}d overdue" if days_ago > 0 else "due today"
                except Exception:
                    when = ""
                lines.append(f"  • {lead['name']} — {lead.get('email') or lead.get('phone') or 'no contact'} ({when})")
            lines.append("")
            lines.append("Say 'mark [name] as followed_up' when done.")
        else:
            lines.append("No follow-ups due right now.")

        # Show last 5 not-contacted
        nc = [l for l in pipeline if l.get("status") == "not_contacted"][:5]
        if nc:
            lines.append("")
            lines.append(f"NEXT UP — NOT CONTACTED YET")
            lines.append("─" * 36)
            for lead in nc:
                lines.append(f"  • {lead['name']} — {lead.get('priority','')} — {lead.get('location','')}")

        return "\n".join(lines)

    # ── Google Places pagination ───────────────────────────────────────────────

    async def _fetch_places(self, niche: str, location: str, count: int, api_key: str) -> list:
        url   = "https://maps.googleapis.com/maps/api/place/textsearch/json"
        query = f"{niche} in {location}"
        places: list = []
        token  = None
        pages_needed = min(3, -(-count // 20))  # ceil(count/20), max 3

        for _ in range(pages_needed):
            params = {"query": query, "key": api_key}
            if token:
                params = {"pagetoken": token, "key": api_key}
                await asyncio.sleep(2)  # Google requires ~2s before using next_page_token

            resp = await asyncio.to_thread(
                requests.get, url, params=params, timeout=12
            )
            data = resp.json()

            if data.get("status") not in ("OK", "ZERO_RESULTS"):
                break

            places.extend(data.get("results", []))
            token = data.get("next_page_token")
            if not token:
                break

        return places

    # ── Enrich one place ──────────────────────────────────────────────────────

    async def _enrich_place(self, place: dict, api_key: str, sem: asyncio.Semaphore = None) -> dict | None:
        async def _do():
            place_id = place.get("place_id")
            if not place_id:
                return None

            # Places Details
            resp = await asyncio.to_thread(
                requests.get,
                "https://maps.googleapis.com/maps/api/place/details/json",
                params={
                    "place_id": place_id,
                    "fields":   ("name,formatted_phone_number,website,formatted_address,"
                                 "rating,user_ratings_total,business_status"),
                    "key":      api_key,
                },
                timeout=12
            )
            d = resp.json().get("result", {})

            if d.get("business_status") == "PERMANENTLY_CLOSED":
                return None

            name    = d.get("name",                   place.get("name", "Unknown"))
            phone   = d.get("formatted_phone_number", "")
            website = d.get("website",                "")
            address = d.get("formatted_address",      "")
            rating  = d.get("rating",                 0) or 0
            reviews = d.get("user_ratings_total",     0) or 0

            # Tidy address
            if address and address.count(",") >= 2:
                parts   = [p.strip() for p in address.split(",")]
                address = ", ".join(parts[:3])

            # ── Site check (LeadForge) ────────────────────────────────────────
            skip       = False
            signals    = []
            email      = ""
            perf_score = None   # PageSpeed mobile 0-100

            if website and _LF_AVAILABLE:
                skip, signals = await asyncio.to_thread(_sitecheck.analyze, website)
                if not skip:
                    email, perf_score = await asyncio.gather(
                        asyncio.to_thread(_scrape.scrape_site, website),
                        self._quick_pagespeed(website),
                    )
            elif not website:
                signals = ["no website"]

            # ── Priority ──────────────────────────────────────────────────────
            if not website:
                priority = "Hot"
                sort_key = 0
            elif skip:
                priority = "Cold"
                sort_key = 2
            else:
                if any(s in signals for s in ["no website", "website unreachable"]):
                    priority = "Hot"
                    sort_key = 0
                elif signals:
                    priority = "Warm"
                    sort_key = 1
                else:
                    priority = "Cold"
                    sort_key = 2

            # ── Build "why" with performance data ─────────────────────────────
            why_parts = [SIGNAL_SENTENCES.get(s, s.title()) for s in signals[:2]]
            if perf_score is not None:
                if perf_score < 50:
                    why_parts.append(f"Very slow site ({perf_score}/100 on mobile)")
                elif perf_score < 80:
                    why_parts.append(f"Slow on mobile ({perf_score}/100 PageSpeed)")
            if why_parts:
                why = " | ".join(why_parts)
            elif skip:
                why = f"Modern site — possible SEO or redesign upsell ({reviews} reviews)"
            else:
                why = f"Established ({reviews} reviews, {rating}★)"

            return {
                "place_id":   place_id,
                "name":       name,
                "phone":      phone,
                "email":      email or "",
                "website":    website,
                "address":    address,
                "rating":     round(float(rating), 1),
                "reviews":    int(reviews),
                "signals":    signals,
                "perf_score": perf_score,
                "priority":   priority,
                "why":        why,
                "sort_key":   sort_key,
            }

        if sem:
            async with sem:
                return await _do()
        return await _do()

    # ── Quick PageSpeed score ─────────────────────────────────────────────────

    async def _quick_pagespeed(self, url: str) -> int | None:
        """Return mobile PageSpeed score (0-100) or None on failure. Fast timeout."""
        try:
            resp = await asyncio.wait_for(
                asyncio.to_thread(
                    requests.get,
                    "https://www.googleapis.com/pagespeedonline/v5/runPagespeed",
                    params={"url": url, "strategy": "mobile", "category": "performance"},
                    timeout=15,
                ),
                timeout=18
            )
            data  = resp.json()
            score = data.get("lighthouseResult", {}).get("categories", {}).get("performance", {}).get("score")
            return int(score * 100) if score is not None else None
        except Exception:
            return None

    # ── Excel spreadsheet ─────────────────────────────────────────────────────

    def _make_spreadsheet(self, leads: list, niche: str, location: str, path: str):
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "4F46E5"

        hot  = sum(1 for l in leads if l["priority"] == "Hot")
        warm = sum(1 for l in leads if l["priority"] == "Warm")
        cold = sum(1 for l in leads if l["priority"] == "Cold")

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color.lstrip("#"))

        def tborder(color="D0D0E0"):
            s = Side(style="thin", color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        def cfont(bold=False, color="1F1F3A", size=10, italic=False):
            return Font(bold=bold, color=color.lstrip("#"), size=size,
                        name="Calibri", italic=italic)

        def calign(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v,
                             wrap_text=wrap, indent=(1 if h == "left" else 0))

        NUM_COLS = 14
        last_col = get_column_letter(NUM_COLS)

        # ── Row 1: Title banner ──
        ws.merge_cells(f"A1:{last_col}1")
        c = ws["A1"]
        c.value     = f"  Lead Report — {niche.title()} in {location.title()}"
        c.font      = Font(bold=True, color="FFFFFF", size=15, name="Calibri")
        c.fill      = fill("1A1A2E")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 42

        # ── Row 2: Stats bar ──
        ws.merge_cells(f"A2:{last_col}2")
        c = ws["A2"]
        c.value = (f"  {len(leads)} businesses  |  "
                   f"{hot} Hot (no website)  |  "
                   f"{warm} Warm (weak site)  |  "
                   f"{cold} Cold (has website)  |  "
                   f"Generated {datetime.now().strftime('%b %d, %Y')} by FileFlow Inc")
        c.font      = Font(color="9999BB", size=9, name="Calibri", italic=True)
        c.fill      = fill("12122A")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 20

        # ── Row 3: Spacer ──
        ws.row_dimensions[3].height = 8

        # ── Row 4: Headers ──
        COLS = [
            ("#",             4.5,  "center"),
            ("Priority",      10,   "center"),
            ("Business Name", 32,   "left"),
            ("Phone",         17,   "left"),
            ("Email",         30,   "left"),
            ("Website",       34,   "left"),
            ("Speed",          9,   "center"),
            ("Address",       30,   "left"),
            ("City / State",  20,   "left"),
            ("Rating",         8,   "center"),
            ("Reviews",        9,   "center"),
            ("Issues Found",  34,   "left"),
            ("Status",        14,   "center"),
            ("Notes",         22,   "left"),
        ]

        for ci, (label, width, _) in enumerate(COLS, 1):
            c = ws.cell(row=4, column=ci, value=label)
            c.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            c.fill      = fill("4F46E5")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = tborder("3730A3")
            ws.column_dimensions[get_column_letter(ci)].width = width

        ws.row_dimensions[4].height = 26
        ws.freeze_panes = "A5"

        ROW_COLORS = {
            "Hot":  ("FEF2F2", "991B1B", "FEE2E2"),
            "Warm": ("FFFBEB", "92400E", "FEF3C7"),
            "Cold": ("F0FDF4", "166534", "DCFCE7"),
        }

        for ri, lead in enumerate(leads, 5):
            is_alt  = (ri % 2 == 0)
            pri     = lead.get("priority", "Cold")
            bg, pfg, alt_bg = ROW_COLORS.get(pri, ("FFFFFF", "374151", "F9FAFB"))
            row_bg  = alt_bg if is_alt else bg

            addr_full = lead.get("address", "")
            if addr_full and "," in addr_full:
                parts     = [p.strip() for p in addr_full.split(",")]
                addr_show = ", ".join(parts[:2])
                city_show = ", ".join(parts[1:3])
            else:
                addr_show = addr_full
                city_show = ""

            website    = lead.get("website",    "") or ""
            rating     = lead.get("rating",     0) or 0
            reviews    = lead.get("reviews",    0) or 0
            email      = lead.get("email",      "") or ""
            perf_score = lead.get("perf_score", None)
            signals    = lead.get("signals",    [])
            issues     = " | ".join(s.title() for s in signals) if signals else ("Good site" if website else "No website")

            # Speed cell value
            if perf_score is not None:
                speed_val = f"{perf_score}/100"
            elif not website:
                speed_val = "N/A"
            else:
                speed_val = "—"

            values = [
                ri - 4,                                      # col 1  #
                pri,                                          # col 2  Priority
                lead.get("name", ""),                        # col 3  Business Name
                lead.get("phone", "") or "",                  # col 4  Phone
                email,                                        # col 5  Email
                website or "NO WEBSITE",                      # col 6  Website
                speed_val,                                    # col 7  Speed
                addr_show or "",                              # col 8  Address
                city_show or "",                              # col 9  City/State
                round(float(rating), 1) if rating else "",   # col 10 Rating
                int(reviews) if reviews else 0,              # col 11 Reviews
                issues,                                       # col 12 Issues Found
                "Not contacted",                              # col 13 Status
                "",                                           # col 14 Notes
            ]

            for ci, (val, (_, _, align_h)) in enumerate(zip(values, COLS), 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border    = tborder()
                c.fill      = fill(row_bg)
                c.font      = cfont(color="1F1F3A")
                c.alignment = calign(h=align_h, wrap=(ci == 12))

                if ci == 2:  # Priority badge
                    c.font = cfont(bold=True, color=pfg)
                    c.fill = fill(alt_bg)

                if ci == 3:  # Business name bold
                    c.font = cfont(bold=True, color="111827")

                if ci == 5 and email:  # Email — blue
                    c.font = cfont(color="1D4ED8")

                if ci == 6 and not website:  # No website — red bold
                    c.font = Font(bold=True, color="DC2626", size=10, name="Calibri")

                # Speed score — color coded
                if ci == 7 and perf_score is not None:
                    if perf_score < 50:
                        c.font = cfont(bold=True, color="DC2626")   # red
                    elif perf_score < 80:
                        c.font = cfont(bold=True, color="D97706")   # amber
                    else:
                        c.font = cfont(bold=True, color="16A34A")   # green

                if ci == 10 and val:  # Rating ★
                    c.value = f"{val} ★"
                    c.alignment = calign("center")

                if ci == 12:  # Issues — smaller muted
                    c.font = cfont(color="4B5563", size=9)

                if ci == 13:  # Status — italic muted
                    c.font = Font(color="9CA3AF", size=9, name="Calibri", italic=True)
                    c.alignment = calign("center")

            ws.row_dimensions[ri].height = 18

        # ── Excel Table ──
        data_last = 4 + len(leads)
        tbl = Table(displayName="LeadTable", ref=f"A4:{last_col}{data_last}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True,  showColumnStripes=False,
        )
        ws.add_table(tbl)

        # ── Summary sheet ──
        ws2 = wb.create_sheet("Summary")
        ws2.sheet_view.showGridLines = False
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 26

        for ri, (k, v, bold, bg, fg, sz) in enumerate([
            ("FRIDAY Lead Report",    None,                              True,  "1A1A2E", "FFFFFF", 13),
            ("",                      None,                              False, "FFFFFF", "1F1F3A", 10),
            ("Search",                f"{niche.title()} in {location.title()}", True, "F5F5FF", "1F1F3A", 10),
            ("Date Generated",        datetime.now().strftime("%B %d, %Y"), False,"FFFFFF", "374151", 10),
            ("Total Leads",           len(leads),                        False, "FFFFFF", "374151", 10),
            ("",                      None,                              False, "FFFFFF", "1F1F3A", 10),
            ("Hot (no website)",      hot,                               True,  "FEF2F2", "991B1B", 10),
            ("Warm (weak site)",      warm,                              True,  "FFFBEB", "92400E", 10),
            ("Cold (has website)",    cold,                              True,  "F0FDF4", "166534", 10),
            ("",                      None,                              False, "FFFFFF", "1F1F3A", 10),
            ("Generated by",          "FRIDAY — FileFlow Inc",           False, "F9F9FF", "6B7280",  9),
        ], 1):
            for ci, val in enumerate([k, v], 1):
                c = ws2.cell(ri, ci, val)
                c.font      = Font(bold=bold, color=fg, size=sz, name="Calibri")
                c.fill      = fill(bg)
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                if k:
                    c.border = tborder()
            ws2.row_dimensions[ri].height = 22

        wb.save(path)

    # ── Email helpers ─────────────────────────────────────────────────────────

    async def _generate_clean_email(self, context: str, to_email: str) -> str:
        prompt = (
            f"Write a cold outreach email for a web design agency (FileFlow) "
            f"targeting: {context}.\n\n"
            f"OUTPUT ONLY this exact format — no extra text, no commentary, nothing else:\n\n"
            f"SUBJECT: [subject line under 8 words]\n"
            f"---\n"
            f"[email body — 4 to 5 sentences, human tone, ends with a soft CTA]\n"
            f"---\n\n"
            f"Rules: no spam words, no corporate speak, first line must be specific to their business."
        )
        if self.provider == "anthropic":
            resp = await asyncio.to_thread(
                self.anthropic_client.messages.create,
                model=self.model,
                max_tokens=600,
                system="You output cold email drafts in the exact format requested. Nothing else.",
                messages=[{"role": "user", "content": prompt}]
            )
            return resp.content[0].text.strip()
        else:
            resp = await asyncio.to_thread(
                self.openai_client.chat.completions.create,
                model="gpt-4o-mini",
                max_tokens=600,
                messages=[
                    {"role": "system", "content": "You output cold email drafts in the exact format requested. Nothing else."},
                    {"role": "user",   "content": prompt}
                ]
            )
            return resp.choices[0].message.content.strip()

    def _parse_email(self, draft: str):
        subject = "Quick question about your website"
        body    = draft
        lines   = draft.split('\n')

        for i, line in enumerate(lines):
            if line.upper().startswith("SUBJECT:"):
                subject = line.split(":", 1)[1].strip()
                rest    = '\n'.join(lines[i + 1:])
                rest    = re.sub(r'^-{3,}\s*$', '', rest, flags=re.MULTILINE)

                for pattern in [
                    r'\nFOLLOW.?UP', r'\nP\.?S\.?', r'\nLet me know',
                    r'\nWant me to', r'\nShould I',  r'\nBoss,?\s*I',
                    r'\n---+',
                ]:
                    m = re.search(pattern, rest, re.IGNORECASE)
                    if m:
                        rest = rest[:m.start()]

                body = rest.strip()
                break

        return subject, body

    def _send_via_brevo(self, to_email: str, subject: str, body: str) -> str:
        if not BREVO_API_KEY:
            return "No Brevo API key in config."
        if not COMPANY_EMAIL:
            return "No sender email in config."

        reply_to = os.getenv("REPLY_TO_EMAIL", COMPANY_EMAIL)
        payload  = {
            "sender":      {"name": COMPANY_NAME, "email": COMPANY_EMAIL},
            "to":          [{"email": to_email}],
            "replyTo":     {"email": reply_to},
            "subject":     subject,
            "textContent": body,
        }
        try:
            resp = requests.post(
                "https://api.brevo.com/v3/smtp/email",
                json=payload,
                headers={"api-key": BREVO_API_KEY, "Content-Type": "application/json"},
                timeout=10,
            )
            if resp.status_code in (200, 201):
                return f"Sent. Email's on its way, Boss."
            return f"Brevo error {resp.status_code}: {resp.text}"
        except Exception as e:
            return f"Failed to send: {str(e)}"
